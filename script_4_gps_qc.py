"""
SCRIPT 4 — GPS & DATA QUALITY CONTROL
================================================
What it does:
  - Checks GPS accuracy per point (flags > threshold)
  - Detects out-of-boundary coordinates
  - Checks if each worker collected exactly 10 points/day
  - Detects duplicate coordinates (same lat/lon submitted twice)
  - Checks timestamp gaps (points collected impossibly fast)
  - Calculates distance between consecutive points
  - Flags copy-pasted points
  - Produces PASS / REVIEW / FAIL per worker per day

Input : cleaned_data.csv
Output: quality_control_report.xlsx (3 sheets)
"""

import sys
import os

# FORCE UTF-8 ENCODING TO PREVENT CHARMAP ERRORS
sys.stdout.reconfigure(encoding='utf-8')

import pandas as pd
import numpy as np
from math import radians, sin, cos, sqrt, atan2
from datetime import datetime
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# ── CONFIG ──────────────────────────────────────────────────────────────
INPUT_FILE          = "cleaned_data.csv"
OUTPUT_FILE         = "quality_control_report.xlsx"

ACCURACY_THRESHOLD  = 10.0      # metres — above this = bad GPS
POINTS_REQUIRED     = 10        # expected points per worker per day
MIN_POINT_GAP_MIN   = 5         # minimum minutes between consecutive points
BOUNDARY = {                    # Bounding box — adjust for your study area
    "lat_min": -1.35,
    "lat_max": -1.25,
    "lon_min": 36.80,
    "lon_max": 36.85,
}
PASS_THRESHOLD    = 95.0        # validity % to PASS
REVIEW_THRESHOLD  = 80.0        # validity % for REVIEW (else FAIL)

# ── HELPERS ──────────────────────────────────────────────────────────────
def thin_border():
    s = Side(style="thin")
    return Border(left=s, right=s, top=s, bottom=s)

def hdr(ws, row=1, color="1F4E79"):
    for cell in ws[row]:
        cell.font = Font(bold=True, color="FFFFFF", name="Calibri", size=10)
        cell.fill = PatternFill("solid", start_color=color)
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = thin_border()
    ws.row_dimensions[row].height = 35

def haversine_m(lat1, lon1, lat2, lon2):
    """Distance between two GPS points in metres."""
    R = 6371000
    try:
        lat1, lon1, lat2, lon2 = map(radians, [lat1, lon1, lat2, lon2])
        dlat = lat2 - lat1
        dlon = lon2 - lon1
        a = sin(dlat/2)**2 + cos(lat1)*cos(lat2)*sin(dlon/2)**2
        return round(R * 2 * atan2(sqrt(a), sqrt(1-a)), 2)
    except:
        return np.nan

# ── 1. LOAD ──────────────────────────────────────────────────────────────
print("[LOAD] Loading cleaned data...")
df = pd.read_csv(INPUT_FILE)
df["date"]     = pd.to_datetime(df["date"]).dt.date
df["latitude"]  = pd.to_numeric(df["latitude"],  errors="coerce")
df["longitude"] = pd.to_numeric(df["longitude"], errors="coerce")
df["accuracy_m"]= pd.to_numeric(df["accuracy_m"],errors="coerce")
print(f"   {len(df)} rows loaded")

# ── 2. POINT-LEVEL QC FLAGS ──────────────────────────────────────────────
print("\n Running point-level quality checks...")

# Flag 1: Bad GPS accuracy
df["flag_bad_accuracy"] = (
    df["accuracy_m"].isna() | (df["accuracy_m"] >= ACCURACY_THRESHOLD)
)

# Flag 2: Out of boundary
df["flag_out_of_bounds"] = ~(
    df["latitude"].between(BOUNDARY["lat_min"], BOUNDARY["lat_max"]) &
    df["longitude"].between(BOUNDARY["lon_min"], BOUNDARY["lon_max"])
)

# Flag 3: Missing GPS entirely
df["flag_missing_gps"] = df["latitude"].isna() | df["longitude"].isna()

# Flag 4: Duplicate coordinates within same worker-day
df["coord_key"] = df["latitude"].astype(str) + "_" + df["longitude"].astype(str)
df["flag_duplicate_coord"] = df.duplicated(subset=["worker_id","date","coord_key"], keep=False)

# Composite: any flag = invalid
df["point_valid"] = ~(
    df["flag_bad_accuracy"] | df["flag_out_of_bounds"] |
    df["flag_missing_gps"]  | df["flag_duplicate_coord"]
)

def invalid_reason(row):
    reasons = []
    if row["flag_missing_gps"]:       reasons.append("Missing GPS")
    if row["flag_bad_accuracy"]:      reasons.append(f"Low Accuracy ({row['accuracy_m']}m)")
    if row["flag_out_of_bounds"]:     reasons.append("Out of Bounds")
    if row["flag_duplicate_coord"]:   reasons.append("Duplicate Coord")
    return " | ".join(reasons) if reasons else ""

df["invalid_reason"] = df.apply(invalid_reason, axis=1)

# ── 3. TIMESTAMP GAP CHECK ───────────────────────────────────────────────
print("  Checking collection timestamps...")

def parse_time(t):
    for fmt in ["%H:%M:%S", "%H:%M"]:
        try:
            return datetime.strptime(str(t).strip(), fmt)
        except:
            pass
    return None

df["collection_dt"] = df["collection_time"].apply(parse_time)
df_sorted = df.sort_values(["worker_id","date","point_number"])

df_sorted["prev_time"]    = df_sorted.groupby(["worker_id","date"])["collection_dt"].shift(1)
df_sorted["time_gap_min"] = df_sorted.apply(
    lambda r: round((r["collection_dt"] - r["prev_time"]).seconds / 60, 1)
    if pd.notna(r["collection_dt"]) and r["prev_time"] is not None else np.nan, axis=1
)
df_sorted["flag_too_fast"] = df_sorted["time_gap_min"] < MIN_POINT_GAP_MIN

# ── 4. DISTANCE BETWEEN POINTS ───────────────────────────────────────────
print(" Calculating point distances...")
df_sorted["prev_lat"] = df_sorted.groupby(["worker_id","date"])["latitude"].shift(1)
df_sorted["prev_lon"] = df_sorted.groupby(["worker_id","date"])["longitude"].shift(1)
df_sorted["dist_from_prev_m"] = df_sorted.apply(
    lambda r: haversine_m(r["prev_lat"], r["prev_lon"], r["latitude"], r["longitude"])
    if pd.notna(r.get("prev_lat")) and pd.notna(r["latitude"]) else np.nan, axis=1
)

# ── 5. PER-POINT REPORT ──────────────────────────────────────────────────
point_report = df_sorted[[
    "worker_id","worker_name","team","date","point_number",
    "latitude","longitude","accuracy_m","altitude_m",
    "point_valid","invalid_reason",
    "collection_time","time_gap_min","flag_too_fast","dist_from_prev_m"
]].copy()
point_report.columns = [
    "Worker ID","Worker Name","Team","Date","Point #",
    "Latitude","Longitude","Accuracy (m)","Altitude (m)",
    "Valid","Invalid Reason",
    "Collection Time","Gap from Prev (min)","Too Fast?","Dist from Prev (m)"
]
point_report["Date"] = point_report["Date"].astype(str)
point_report["Valid"] = point_report["Valid"].map({True:"[DONE] YES", False:"[ERROR] NO"})
point_report["Too Fast?"] = point_report["Too Fast?"].map({True:"[WARN] YES", False:"OK", np.nan: ""})


# ── 6. DAILY QC SUMMARY ─────────────────────────────────────────────────
print(" Building daily QC summary...")
daily_qc = df_sorted.groupby(["worker_id","worker_name","team","date"]).agg(
    points_submitted   = ("point_number",        "count"),
    valid_points       = ("point_valid",          "sum"),
    bad_accuracy       = ("flag_bad_accuracy",    "sum"),
    out_of_bounds      = ("flag_out_of_bounds",   "sum"),
    duplicate_coords   = ("flag_duplicate_coord", "sum"),
    too_fast_points    = ("flag_too_fast",         "sum"),
    avg_accuracy_m     = ("accuracy_m",           "mean"),
    avg_dist_m         = ("dist_from_prev_m",     "mean"),
).reset_index()

daily_qc["invalid_points"]   = daily_qc["points_submitted"] - daily_qc["valid_points"]
daily_qc["points_missing"]   = POINTS_REQUIRED - daily_qc["points_submitted"]
daily_qc["points_missing"]   = daily_qc["points_missing"].clip(lower=0)
daily_qc["validity_pct"]     = round(daily_qc["valid_points"] / daily_qc["points_submitted"] * 100, 1)
daily_qc["avg_accuracy_m"]   = daily_qc["avg_accuracy_m"].round(1)
daily_qc["avg_dist_m"]       = daily_qc["avg_dist_m"].round(1)

def qc_status(row):
    if row["validity_pct"] >= PASS_THRESHOLD and row["points_missing"] == 0:
        return "PASS"
    elif row["validity_pct"] >= REVIEW_THRESHOLD:
        return "REVIEW"
    else:
        return "FAIL"

daily_qc["qc_status"] = daily_qc.apply(qc_status, axis=1)
daily_qc["date"]      = daily_qc["date"].astype(str)

# ── 7. WORKER QC SUMMARY ────────────────────────────────────────────────
worker_qc = df_sorted.groupby(["worker_id","worker_name","team"]).agg(
    total_points    = ("point_number",        "count"),
    valid_points    = ("point_valid",          "sum"),
    bad_accuracy    = ("flag_bad_accuracy",    "sum"),
    out_of_bounds   = ("flag_out_of_bounds",   "sum"),
    dup_coords      = ("flag_duplicate_coord", "sum"),
    too_fast        = ("flag_too_fast",         "sum"),
).reset_index()
worker_qc["invalid_points"] = worker_qc["total_points"] - worker_qc["valid_points"]
worker_qc["validity_pct"]   = round(worker_qc["valid_points"] / worker_qc["total_points"] * 100, 1)
worker_qc["qc_status"]      = worker_qc["validity_pct"].apply(
    lambda v: "PASS" if v >= PASS_THRESHOLD else ("REVIEW" if v >= REVIEW_THRESHOLD else "FAIL")
)

# ── 8. WRITE EXCEL ───────────────────────────────────────────────────────
print(f"\n[SAVE] Writing {OUTPUT_FILE}...")
STATUS_FILL = {"PASS":"D5F5E3", "REVIEW":"FEF9E7", "FAIL":"FADBD8"}

with pd.ExcelWriter(OUTPUT_FILE, engine="openpyxl") as writer:

    # Sheet 1: Worker QC Summary
    worker_qc.to_excel(writer, sheet_name="Worker QC Summary", index=False)
    ws1 = writer.sheets["Worker QC Summary"]
    hdr(ws1)
    status_col = list(worker_qc.columns).index("qc_status") + 1
    for row_idx in range(2, len(worker_qc)+2):
        status = str(ws1.cell(row=row_idx, column=status_col).value or "")
        fill = STATUS_FILL.get(status, "FFFFFF")
        for col_idx in range(1, len(worker_qc.columns)+1):
            cell = ws1.cell(row=row_idx, column=col_idx)
            cell.fill = PatternFill("solid", start_color=fill)
            cell.font = Font(name="Calibri", size=10,
                             bold=(col_idx == status_col))
            cell.border = thin_border()
            cell.alignment = Alignment(horizontal="center")
    for col_idx, col in enumerate(worker_qc.columns, 1):
        ws1.column_dimensions[get_column_letter(col_idx)].width = max(len(str(col))+2, 14)
    ws1.freeze_panes = "A2"

    # Sheet 2: Daily QC
    daily_qc.to_excel(writer, sheet_name="Daily QC", index=False)
    ws2 = writer.sheets["Daily QC"]
    hdr(ws2, color="2E4057")
    sc2 = list(daily_qc.columns).index("qc_status") + 1
    for row_idx in range(2, len(daily_qc)+2):
        status = str(ws2.cell(row=row_idx, column=sc2).value or "")
        fill = STATUS_FILL.get(status, "FFFFFF")
        for col_idx in range(1, len(daily_qc.columns)+1):
            cell = ws2.cell(row=row_idx, column=col_idx)
            cell.fill = PatternFill("solid", start_color=fill)
            cell.font = Font(name="Calibri", size=9)
            cell.border = thin_border()
            cell.alignment = Alignment(horizontal="center")
    for col_idx, col in enumerate(daily_qc.columns, 1):
        ws2.column_dimensions[get_column_letter(col_idx)].width = max(len(str(col))+2, 12)
    ws2.freeze_panes = "A2"

    # Sheet 3: Point-Level Report
    point_report.to_excel(writer, sheet_name="Point Level Report", index=False)
    ws3 = writer.sheets["Point Level Report"]
    hdr(ws3, color="5D3A9B")
    valid_col = list(point_report.columns).index("Valid") + 1
    for row_idx in range(2, len(point_report)+2):
        valid = str(ws3.cell(row=row_idx, column=valid_col).value or "")
        fill = "D5F5E3" if "YES" in valid else ("FADBD8" if "NO" in valid else "FFFFFF")
        for col_idx in range(1, len(point_report.columns)+1):
            cell = ws3.cell(row=row_idx, column=col_idx)
            cell.fill = PatternFill("solid", start_color=fill)
            cell.font = Font(name="Calibri", size=9)
            cell.border = thin_border()
            cell.alignment = Alignment(horizontal="center")
    for col_idx, col in enumerate(point_report.columns, 1):
        ws3.column_dimensions[get_column_letter(col_idx)].width = max(len(str(col))+2, 12)
    ws3.freeze_panes = "A2"

# Summary print
passes  = (worker_qc["qc_status"] == "PASS").sum()
reviews = (worker_qc["qc_status"] == "REVIEW").sum()
fails   = (worker_qc["qc_status"] == "FAIL").sum()
total   = len(df)
valid   = df["point_valid"].sum()


print(f"\n[DONE] DONE")
print(f"   Total points:   {total}")
print(f"   Valid points:   {valid} ({round(valid/total*100,1)}%)")
print(f"   Invalid points: {total - valid}")
# REPLACED UNICODE ARROW WITH '->'
print(f"\n   Worker QC Results -> PASS: {passes} | REVIEW: {reviews} | FAIL: {fails}")