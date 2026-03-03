"""
SCRIPT 1 — ODK DATA CLEANING
================================================
What it does:
  - Loads your raw ODK Excel export
  - Drops redundant/hidden ODK system columns
  - Renames ugly ODK column names to clean ones
  - Standardizes date, time, and text formats
  - Flags and fills missing values
  - Saves a clean version ready for all other scripts

Input : ODK_Raw_Export.xlsx
Output: cleaned_data.xlsx + cleaned_data.csv
"""

import pandas as pd
import numpy as np
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# ── CONFIG ──────────────────────────────────────────────────────────────
INPUT_FILE  = "ODK_Raw_Export.xlsx"
OUTPUT_XLSX = "cleaned_data.xlsx"
OUTPUT_CSV  = "cleaned_data.csv"

# ── 1. LOAD ──────────────────────────────────────────────────────────────
print("[LOAD] Loading ODK raw data...")
df = pd.read_excel(INPUT_FILE, dtype=str)
print(f"   Loaded {len(df)} rows × {len(df.columns)} columns")

# ── 2. DROP REDUNDANT ODK SYSTEM COLUMNS ────────────────────────────────
print("\n[CLEAN]  Dropping redundant ODK system columns...")
drop_cols = [
    "subscriberid", "simserial", "KEY", "_version", "_edited",
    "formhub/uuid", "meta/instanceID",
    # Hidden duplicate GPS columns ODK adds
    "_gps_location_latitude", "_gps_location_longitude",
    "_gps_location_altitude", "_gps_location_precision",
    # second point duplicates if present
]
drop_cols = [c for c in drop_cols if c in df.columns]
df.drop(columns=drop_cols, inplace=True)
#print(f"   Dropped {len(drop_cols)} columns → {len(df.columns)} remaining")
import sys

# This forces the output to handle characters your terminal might not support
sys.stdout.reconfigure(encoding='utf-8')

print(f"    Dropped {len(drop_cols)} columns \u2192 {len(df.columns)} remaining")

#print(f"    Dropped {len(drop_cols)} columns -> {len(df.columns)} remaining")

# ── 3. RENAME COLUMNS ───────────────────────────────────────────────────
print("\n[RENAME]  Renaming columns to clean names...")
rename_map = {
    "start"                    : "session_start",
    "end"                      : "session_end",
    "today"                    : "date",
    "deviceid"                 : "device_id",
    "phonenumber"              : "phone_number",
    "username"                 : "odk_username",
    "instanceID"               : "instance_id",
    "SubmissionDate"           : "submission_date",
    "_id"                      : "record_id",
    "_uuid"                    : "uuid",
    "_submission_time"         : "submission_time",
    "_duration"                : "duration_seconds",
    "_status"                  : "submission_status",
    "gps_location-Latitude"    : "latitude",
    "gps_location-Longitude"   : "longitude",
    "gps_location-Altitude"    : "altitude_m",
    "gps_location-Accuracy"    : "accuracy_m",
    "second_point-Latitude"    : "second_lat",
    "second_point-Longitude"   : "second_lon",
    "second_point-Altitude"    : "second_alt",
    "second_point-Accuracy"    : "second_accuracy_m",
    "vegetation_cover"         : "vegetation",
    "water_presence"           : "water",
    "canopy_cover_pct"         : "canopy_pct",
    "photo_taken"              : "photo",
}
rename_map = {k: v for k, v in rename_map.items() if k in df.columns}
df.rename(columns=rename_map, inplace=True)
print(f"   Renamed {len(rename_map)} columns")

# ── 4. FIX DATA TYPES ───────────────────────────────────────────────────
print("\n[PROCESS] Fixing data types...")

# Numeric
for col in ["latitude","longitude","altitude_m","accuracy_m",
            "second_lat","second_lon","second_alt","second_accuracy_m",
            "canopy_pct","duration_seconds","record_id","point_number"]:
    if col in df.columns:
        df[col] = pd.to_numeric(df[col], errors="coerce")

# Dates
# Dates
for col in ["date","session_start","session_end","submission_date","submission_time"]:
    if col in df.columns:
        df[col] = pd.to_datetime(df[col], errors="coerce")

# 🔥 REMOVE TIMEZONE (Excel cannot handle timezone-aware datetimes)
for col in df.select_dtypes(include=["datetimetz"]).columns:
    df[col] = df[col].dt.tz_localize(None)

# Extract date only where needed
if "date" in df.columns:
    df["date"] = df["date"].dt.date

# Times — keep as string HH:MM:SS
for col in ["clock_in_time","clock_out_time","collection_time"]:
    if col in df.columns:
        df[col] = df[col].astype(str).str.strip().replace("nan","")

# Uppercase categorical fields
for col in ["land_use","vegetation","water","soil_type","slope","photo","submission_status"]:
    if col in df.columns:
        df[col] = df[col].astype(str).str.strip().str.title().replace("Nan","")

# ── 5. STANDARDIZE WORKER FIELDS ────────────────────────────────────────
print("\n👤 Standardizing worker fields...")
if "worker_name" in df.columns:
    df["worker_name"] = df["worker_name"].astype(str).str.strip().str.title()
if "team" in df.columns:
    df["team"] = df["team"].astype(str).str.strip().str.title()

# ── 6. FLAG MISSING VALUES ───────────────────────────────────────────────
print("\n[CHECK] Flagging missing values...")
critical_cols = ["latitude","longitude","accuracy_m","clock_in_time","worker_id"]
for col in critical_cols:
    if col in df.columns:
        missing = df[col].isna().sum()
        if missing > 0:
            print(f"   [WARN]  {col}: {missing} missing values")

# Add a missing-data flag column
df["has_missing_gps"] = (
    df["latitude"].isna() | df["longitude"].isna() | df["accuracy_m"].isna()
).map({True: "YES", False: "NO"})

# Fill non-critical missing with sensible defaults
if "notes" in df.columns:
    df["notes"] = df["notes"].fillna("")
if "canopy_pct" in df.columns:
    df["canopy_pct"] = df["canopy_pct"].fillna(-1)  # -1 = not recorded

# ── 7. ADD COMPUTED HELPER COLUMNS ──────────────────────────────────────
print("\n[COMPUTE] Adding computed columns...")

# Hours worked
def calc_hours(row):
    try:
        ci = pd.to_datetime(str(row.get("clock_in_time","")).strip(), format="%H:%M:%S")
        co = pd.to_datetime(str(row.get("clock_out_time","")).strip(), format="%H:%M:%S")
        return round((co - ci).seconds / 3600, 2)
    except:
        return np.nan

df["hours_worked"] = df.apply(calc_hours, axis=1)

# Clock-in status
def clock_status(row):
    try:
        ci = pd.to_datetime(str(row.get("clock_in_time","")).strip(), format="%H:%M:%S")
        required = pd.to_datetime("08:00:00", format="%H:%M:%S")
        diff = (ci - required).seconds // 60
        if ci < required:
            return f"Early ({abs(int((ci-required).total_seconds()//60))} min early)"
        elif diff <= 5:
            return "On Time"
        else:
            return f"Late ({diff} min late)"
    except:
        return "Unknown"

df["clock_in_status"] = df.apply(clock_status, axis=1)

# ── 8. SORT ─────────────────────────────────────────────────────────────
sort_cols = [c for c in ["worker_id","date","point_number"] if c in df.columns]
df.sort_values(sort_cols, inplace=True)
df.reset_index(drop=True, inplace=True)

# ── 9. SAVE CSV ──────────────────────────────────────────────────────────
df.to_csv(OUTPUT_CSV, index=False)
print(f"\n[SAVE] CSV saved → {OUTPUT_CSV}")

# ── 10. SAVE FORMATTED EXCEL ─────────────────────────────────────────────
print(f"[SAVE] Writing Excel → {OUTPUT_XLSX}")

with pd.ExcelWriter(OUTPUT_XLSX, engine="openpyxl") as writer:
    df.to_excel(writer, sheet_name="Cleaned Data", index=False)
    ws = writer.sheets["Cleaned Data"]

    # Header styling
    hdr_fill = PatternFill("solid", start_color="1F4E79")
    for col_idx, col_name in enumerate(df.columns, 1):
        cell = ws.cell(row=1, column=col_idx)
        cell.font = Font(bold=True, color="FFFFFF", name="Calibri", size=10)
        cell.fill = hdr_fill
        cell.alignment = Alignment(horizontal="center", wrap_text=True)

    # Auto-width
    for col_idx, col_name in enumerate(df.columns, 1):
        max_len = max(len(str(col_name)), df[col_name].astype(str).str.len().max())
        ws.column_dimensions[get_column_letter(col_idx)].width = min(max_len + 2, 30)

    ws.row_dimensions[1].height = 35
    ws.freeze_panes = "A2"

print(f"\n[DONE] DONE — {len(df)} clean rows × {len(df.columns)} columns")
print(f"   Output: {OUTPUT_XLSX}  |  {OUTPUT_CSV}")
