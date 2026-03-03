"""
SCRIPT 2 — ATTENDANCE & CLOCK-IN ANALYSIS
================================================
What it does:
  - Reads cleaned_data.csv (from Script 1)
  - Determines who worked each day, who was absent
  - Calculates minutes early/late per worker per day
  - Builds a daily attendance matrix (like a register)
  - Produces a summary: attendance rate, late days, early days
  - Flags chronic latecomers

Input : cleaned_data.csv
Output: attendance_report.xlsx (3 sheets)
"""
import sys
import os

# FORCE UTF-8 ENCODING TO PREVENT CHARMAP ERRORS
sys.stdout.reconfigure(encoding='utf-8')


import pandas as pd
import numpy as np
from datetime import datetime, time
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# ── CONFIG ──────────────────────────────────────────────────────────────
INPUT_FILE    = "cleaned_data.csv"
OUTPUT_FILE   = "attendance_report.xlsx"
REQUIRED_START = "08:00:00"
REQUIRED_END   = "17:00:00"
LATE_THRESHOLD = 5          # minutes grace period before marked Late
CHRONIC_LATE   = 3          # days late = flagged as chronic

# ── HELPERS ──────────────────────────────────────────────────────────────
def thin_border():
    s = Side(style="thin")
    return Border(left=s, right=s, top=s, bottom=s)

def style_header(ws, row=1, color="1F4E79"):
    for cell in ws[row]:
        cell.font = Font(bold=True, color="FFFFFF", name="Calibri", size=10)
        cell.fill = PatternFill("solid", start_color=color)
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = thin_border()
    ws.row_dimensions[row].height = 35

def auto_width(ws, df):
    for col_idx, col_name in enumerate(df.columns, 1):
        max_len = max(len(str(col_name)), df[col_name].astype(str).str.len().max())
        ws.column_dimensions[get_column_letter(col_idx)].width = min(max_len + 3, 35)

# ── 1. LOAD ──────────────────────────────────────────────────────────────
print("[LOAD] Loading cleaned data...")
df = pd.read_csv(INPUT_FILE)
df["date"] = pd.to_datetime(df["date"]).dt.date
print(f"   {len(df)} rows loaded")

# ── 2. BUILD DAILY ATTENDANCE ────────────────────────────────────────────
print("\n Building daily attendance records...")

def parse_time(t_str):
    for fmt in ["%H:%M:%S", "%H:%M"]:
        try:
            return datetime.strptime(str(t_str).strip(), fmt).time()
        except:
            pass
    return None

required_ci = parse_time(REQUIRED_START)
required_co = parse_time(REQUIRED_END)

# Get one row per worker per day (first submission = clock-in record)
daily = df.drop_duplicates(subset=["worker_id","date"]).copy()

daily["ci_time"]  = daily["clock_in_time"].apply(parse_time)
daily["co_time"]  = daily["clock_out_time"].apply(parse_time)

def minutes_diff(actual, required):
    if actual is None or required is None:
        return np.nan
    a = actual.hour * 60 + actual.minute
    r = required.hour * 60 + required.minute
    return a - r   # negative = early, positive = late

daily["ci_diff_min"] = daily["ci_time"].apply(lambda t: minutes_diff(t, required_ci))
daily["co_diff_min"] = daily["co_time"].apply(lambda t: minutes_diff(t, required_co))

def attendance_status(diff):
    if pd.isna(diff):
        return "Unknown"
    elif diff < -LATE_THRESHOLD:
        return f"Early ({abs(int(diff))} min)"
    elif diff <= LATE_THRESHOLD:
        return "On Time"
    else:
        return f"Late ({int(diff)} min)"

daily["clock_in_status"]  = daily["ci_diff_min"].apply(attendance_status)
daily["clock_out_status"] = daily["co_diff_min"].apply(lambda d: 
    "Left Early" if (not pd.isna(d) and d < -LATE_THRESHOLD) else 
    ("Overtime"  if (not pd.isna(d) and d > 30) else "Normal"))

def calc_hours(row):
    try:
        ci = datetime.combine(datetime.today(), row["ci_time"])
        co = datetime.combine(datetime.today(), row["co_time"])
        hrs = (co - ci).seconds / 3600
        return round(hrs, 2)
    except:
        return np.nan

daily["hours_worked"] = daily.apply(calc_hours, axis=1)

# ── 3. GET ALL EXPECTED WORKER-DAY COMBINATIONS ──────────────────────────
all_workers   = df[["worker_id","worker_name","team"]].drop_duplicates()
all_dates     = sorted(df["date"].unique())
work_days     = [d for d in all_dates if pd.Timestamp(d).weekday() < 5]

expected = pd.MultiIndex.from_product(
    [all_workers["worker_id"].unique(), work_days],
    names=["worker_id","date"]
)
expected_df = pd.DataFrame(index=expected).reset_index()
expected_df = expected_df.merge(all_workers, on="worker_id", how="left")

# Merge actual attendance
att = expected_df.merge(
    daily[["worker_id","date","ci_diff_min","co_diff_min",
           "clock_in_status","clock_out_status","hours_worked",
           "clock_in_time","clock_out_time"]],
    on=["worker_id","date"], how="left"
)
att["present"] = att["clock_in_time"].notna().map({True:"Present", False:"Absent"})

# ── 4. DAILY REGISTER (Matrix View) ─────────────────────────────────────

print(" Building attendance matrix...")
matrix = att.pivot_table(
    index=["worker_id","worker_name","team"],
    columns="date",
    values="present",
    aggfunc="first"
).reset_index()


matrix.columns.name = None
matrix.columns = [str(c) for c in matrix.columns]
matrix = matrix.fillna("Absent")
# ── 5. WORKER SUMMARY ────────────────────────────────────────────────────
print(" Calculating worker summary...")
summary = att.groupby(["worker_id","worker_name","team"]).agg(
    total_work_days    = ("date",    "count"),
    days_present       = ("present", lambda x: (x=="Present").sum()),
    days_absent        = ("present", lambda x: (x=="Absent").sum()),
    days_on_time       = ("clock_in_status", lambda x: x.str.startswith("On Time").sum()),
    days_early         = ("clock_in_status", lambda x: x.str.startswith("Early").sum()),
    days_late          = ("clock_in_status", lambda x: x.str.startswith("Late").sum()),
    avg_hours_per_day  = ("hours_worked", "mean"),
    total_hours        = ("hours_worked", "sum"),
    avg_ci_diff_min    = ("ci_diff_min",  "mean"),
).reset_index()

summary["attendance_rate_pct"] = round(summary["days_present"] / summary["total_work_days"] * 100, 1)
summary["avg_hours_per_day"]   = round(summary["avg_hours_per_day"], 2)
summary["total_hours"]         = round(summary["total_hours"], 2)
summary["avg_ci_diff_min"]     = round(summary["avg_ci_diff_min"], 1)

# Flag chronic latecomers
summary["chronic_late_flag"] = summary["days_late"].apply(
    lambda x: "[WARN] CHRONIC LATE" if x >= CHRONIC_LATE else ""
)

# Sort by attendance rate ascending (worst first)
summary.sort_values("attendance_rate_pct", inplace=True)

# ── 6. DAILY DETAIL TABLE ────────────────────────────────────────────────
detail_cols = ["worker_id","worker_name","team","date","present",
               "clock_in_time","clock_out_time","clock_in_status",
               "clock_out_status","hours_worked","ci_diff_min"]
detail = att[detail_cols].copy()
detail.columns = ["Worker ID","Worker Name","Team","Date","Status",
                  "Clock In","Clock Out","Clock In Status",
                  "Clock Out Status","Hours Worked","Min Late/Early"]
detail["Date"] = detail["Date"].astype(str)

# ── 7. WRITE EXCEL ───────────────────────────────────────────────────────
print(f"\n[SAVE] Writing {OUTPUT_FILE}...")

STATUS_COLORS = {
    "Present" : "D5F5E3",
    "Absent"  : "FADBD8",
    "On Time" : "D5F5E3",
    "Early"   : "D6EAF8",
    "Late"    : "FADBD8",
}

with pd.ExcelWriter(OUTPUT_FILE, engine="openpyxl") as writer:

    # ── Sheet 1: Worker Summary ──
    summary.to_excel(writer, sheet_name="Worker Summary", index=False)
    ws1 = writer.sheets["Worker Summary"]
    style_header(ws1)
    for row_idx in range(2, len(summary)+2):
        flag = ws1.cell(row=row_idx, column=summary.columns.get_loc("chronic_late_flag")+1).value
        for col_idx in range(1, len(summary.columns)+1):
            cell = ws1.cell(row=row_idx, column=col_idx)
            if flag and "CHRONIC" in str(flag):
                cell.fill = PatternFill("solid", start_color="FADBD8")
            elif row_idx % 2 == 0:
                cell.fill = PatternFill("solid", start_color="EBF5FB")
            cell.font = Font(name="Calibri", size=10)
            cell.border = thin_border()
            cell.alignment = Alignment(horizontal="center")
    auto_width(ws1, summary)
    ws1.freeze_panes = "A2"

    # ── Sheet 2: Daily Detail ──
    detail.to_excel(writer, sheet_name="Daily Detail", index=False)
    ws2 = writer.sheets["Daily Detail"]
    style_header(ws2)
    status_col = list(detail.columns).index("Status") + 1
    ci_col     = list(detail.columns).index("Clock In Status") + 1
    for row_idx in range(2, len(detail)+2):
        status = ws2.cell(row=row_idx, column=status_col).value
        ci_st  = str(ws2.cell(row=row_idx, column=ci_col).value or "")
        for col_idx in range(1, len(detail.columns)+1):
            cell = ws2.cell(row=row_idx, column=col_idx)
            if status == "Absent":
                cell.fill = PatternFill("solid", start_color="FADBD8")
            elif ci_st.startswith("Late"):
                cell.fill = PatternFill("solid", start_color="FDEBD0")
            elif row_idx % 2 == 0:
                cell.fill = PatternFill("solid", start_color="EBF5FB")
            cell.font = Font(name="Calibri", size=9)
            cell.border = thin_border()
            cell.alignment = Alignment(horizontal="center")
    auto_width(ws2, detail)
    ws2.freeze_panes = "A2"

    # ── Sheet 3: Attendance Matrix ──
    matrix.to_excel(writer, sheet_name="Attendance Matrix", index=False)
    ws3 = writer.sheets["Attendance Matrix"]
    style_header(ws3, color="2E4057")
    for row_idx in range(2, len(matrix)+2):
        for col_idx in range(1, len(matrix.columns)+1):
            cell = ws3.cell(row=row_idx, column=col_idx)
            val = str(cell.value or "")
            if val == "Present":
                cell.fill = PatternFill("solid", start_color="D5F5E3")
            elif val == "Absent":
                cell.fill = PatternFill("solid", start_color="FADBD8")
            cell.font = Font(name="Calibri", size=9, bold=(col_idx<=3))
            cell.border = thin_border()
            cell.alignment = Alignment(horizontal="center")
    for col_idx in range(1, len(matrix.columns)+1):
        ws3.column_dimensions[get_column_letter(col_idx)].width = 12 if col_idx > 3 else 18
    ws3.freeze_panes = "D2"


print(f"\n[DONE] DONE")
# REPLACED UNICODE ARROWS WITH STANDARD '->'
print(f"   Sheet 1 -> Worker Summary    ({len(summary)} workers)")
print(f"   Sheet 2 -> Daily Detail      ({len(detail)} worker-day records)")
print(f"   Sheet 3 -> Attendance Matrix")
print(f"\n Chronic latecomers flagged: {(summary['days_late'] >= CHRONIC_LATE).sum()}")
