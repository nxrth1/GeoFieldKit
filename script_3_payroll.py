"""
SCRIPT 3 — PAYROLL CALCULATION
================================================
What it does:
  - Reads cleaned_data.csv
  - Calculates daily pay per worker (hours × rate)
  - Applies overtime rules (>8 hrs = 1.5× rate)
  - Handles absent days (no pay) and half days
  - Builds a full payroll register per worker
  - Produces payslip-style summary per worker

Input : cleaned_data.csv
Output: payroll_report.xlsx (3 sheets)
"""

import pandas as pd
import numpy as np
from datetime import datetime
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# ── CONFIG ──────────────────────────────────────────────────────────────
INPUT_FILE      = "cleaned_data.csv"
OUTPUT_FILE     = "payroll_report.xlsx"

HOURLY_RATES = {          # KES per hour per worker (edit to match your real rates)
    "W001": 18.00,
    "W002": 18.00,
    "W003": 18.00,
    "W004": 20.00,
    "W005": 18.00,
    "W006": 20.00,
    "W007": 18.00,
    "W008": 18.00,
    "W009": 20.00,
    "W010": 18.00,
}
DEFAULT_RATE     = 18.00
STANDARD_HOURS   = 8.0       # hours before overtime kicks in
OVERTIME_MULT    = 1.5       # overtime multiplier
CURRENCY         = "KES"

# ── HELPERS ──────────────────────────────────────────────────────────────
def thin_border():
    s = Side(style="thin")
    return Border(left=s, right=s, top=s, bottom=s)

def hdr(ws, row=1, color="1F4E79"):
    for cell in ws[row]:
        cell.font = Font(bold=True, color="FFFFFF", name="Calibri", size=10)
        cell.fill = PatternFill("solid", start_color=color)
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = thin_border()
    ws.row_dimensions[row].height = 35

def parse_time(t_str):
    for fmt in ["%H:%M:%S", "%H:%M"]:
        try:
            return datetime.strptime(str(t_str).strip(), fmt)
        except:
            pass
    return None

def calc_hours(row):
    ci = parse_time(row.get("clock_in_time",""))
    co = parse_time(row.get("clock_out_time",""))
    if ci and co:
        return round((co - ci).seconds / 3600, 2)
    return 0.0

def calc_pay(hours, rate):
    if hours <= 0:
        return 0.0
    if hours <= STANDARD_HOURS:
        return round(hours * rate, 2)
    else:
        regular  = STANDARD_HOURS * rate
        overtime = (hours - STANDARD_HOURS) * rate * OVERTIME_MULT
        return round(regular + overtime, 2)

# ── 1. LOAD ──────────────────────────────────────────────────────────────
print("[LOAD] Loading cleaned data...")
df = pd.read_csv(INPUT_FILE)
df["date"] = pd.to_datetime(df["date"]).dt.date

# ── 2. DAILY HOURS PER WORKER ────────────────────────────────────────────
print("  Calculating daily hours per worker...")
daily = df.drop_duplicates(subset=["worker_id","date"]).copy()
daily["hours_worked"] = daily.apply(calc_hours, axis=1)

# Map hourly rates
daily["hourly_rate"] = daily["worker_id"].map(HOURLY_RATES).fillna(DEFAULT_RATE)

# Calculate regular and overtime hours
daily["regular_hours"]  = daily["hours_worked"].apply(lambda h: min(h, STANDARD_HOURS))
daily["overtime_hours"] = daily["hours_worked"].apply(lambda h: max(0, h - STANDARD_HOURS))

# Calculate pay components
daily["regular_pay"]   = daily.apply(
    lambda r: round(r["regular_hours"]  * r["hourly_rate"], 2), axis=1)
daily["overtime_pay"]  = daily.apply(
    lambda r: round(r["overtime_hours"] * r["hourly_rate"] * OVERTIME_MULT, 2), axis=1)
daily["total_day_pay"] = daily["regular_pay"] + daily["overtime_pay"]

# Day classification
def classify_day(hours):
    if hours == 0:
        return "Absent"
    elif hours < 4:
        return "Partial (<4h)"
    elif hours < STANDARD_HOURS:
        return "Half Day"
    elif hours == STANDARD_HOURS:
        return "Full Day"
    else:
        return f"Overtime (+{round(hours-STANDARD_HOURS,1)}h)"

daily["day_type"] = daily["hours_worked"].apply(classify_day)

# ── 3. DAILY PAYROLL REGISTER ────────────────────────────────────────────
register = daily[[
    "worker_id","worker_name","team","date",
    "clock_in_time","clock_out_time","hours_worked",
    "regular_hours","overtime_hours","hourly_rate",
    "regular_pay","overtime_pay","total_day_pay","day_type"
]].copy()
register.columns = [
    "Worker ID","Worker Name","Team","Date",
    "Clock In","Clock Out","Hours Worked",
    "Regular Hrs","Overtime Hrs","Hourly Rate ("+CURRENCY+")",
    "Regular Pay","Overtime Pay","Total Day Pay","Day Type"
]
register["Date"] = register["Date"].astype(str)
register = register.sort_values(["Worker ID","Date"])

# ── 4. WORKER PAYROLL SUMMARY ────────────────────────────────────────────
print(" Summarising payroll by worker...")
pay_summary = daily.groupby(["worker_id","worker_name","team"]).agg(
    hourly_rate       = ("hourly_rate",    "first"),
    days_worked       = ("hours_worked",   lambda x: (x > 0).sum()),
    days_absent       = ("hours_worked",   lambda x: (x == 0).sum()),
    total_hours       = ("hours_worked",   "sum"),
    regular_hours     = ("regular_hours",  "sum"),
    overtime_hours    = ("overtime_hours", "sum"),
    regular_pay       = ("regular_pay",    "sum"),
    overtime_pay      = ("overtime_pay",   "sum"),
    total_pay         = ("total_day_pay",  "sum"),
).reset_index()

pay_summary["total_hours"]    = pay_summary["total_hours"].round(2)
pay_summary["regular_hours"]  = pay_summary["regular_hours"].round(2)
pay_summary["overtime_hours"] = pay_summary["overtime_hours"].round(2)
pay_summary["regular_pay"]    = pay_summary["regular_pay"].round(2)
pay_summary["overtime_pay"]   = pay_summary["overtime_pay"].round(2)
pay_summary["total_pay"]      = pay_summary["total_pay"].round(2)

# Totals row
totals = pay_summary[[
    "days_worked","days_absent","total_hours",
    "regular_hours","overtime_hours","regular_pay",
    "overtime_pay","total_pay"
]].sum()
totals_row = pd.DataFrame([{
    "worker_id":"—","worker_name":"GRAND TOTAL","team":"—",
    "hourly_rate":"—",
    **totals.to_dict()
}])
pay_summary_display = pd.concat([pay_summary, totals_row], ignore_index=True)

# ── 5. MONTHLY PAYROLL OVERVIEW ──────────────────────────────────────────
daily["month"] = pd.to_datetime(daily["date"].astype(str)).dt.strftime("%Y-%m")
monthly = daily.groupby(["worker_id","worker_name","month"]).agg(
    days_worked   = ("hours_worked",  lambda x: (x > 0).sum()),
    total_hours   = ("hours_worked",  "sum"),
    total_pay     = ("total_day_pay", "sum"),
).reset_index()
monthly["total_hours"] = monthly["total_hours"].round(2)
monthly["total_pay"]   = monthly["total_pay"].round(2)

# ── 6. WRITE EXCEL ───────────────────────────────────────────────────────
print(f"[SAVE] Writing {OUTPUT_FILE}...")
with pd.ExcelWriter(OUTPUT_FILE, engine="openpyxl") as writer:

    # Sheet 1: Payroll Summary
    pay_summary_display.to_excel(writer, sheet_name="Payroll Summary", index=False)
    ws1 = writer.sheets["Payroll Summary"]
    hdr(ws1)
    last_data_row = len(pay_summary_display) + 1
    for row_idx in range(2, last_data_row + 1):
        is_total = (row_idx == last_data_row)
        for col_idx in range(1, len(pay_summary_display.columns)+1):
            cell = ws1.cell(row=row_idx, column=col_idx)
            if is_total:
                cell.font = Font(bold=True, color="FFFFFF", name="Calibri", size=10)
                cell.fill = PatternFill("solid", start_color="1F4E79")
            elif row_idx % 2 == 0:
                cell.fill = PatternFill("solid", start_color="EBF5FB")
                cell.font = Font(name="Calibri", size=10)
            else:
                cell.font = Font(name="Calibri", size=10)
            cell.border = thin_border()
            cell.alignment = Alignment(horizontal="center")
            # Currency format
            if col_idx >= 10:
                cell.number_format = f'#,##0.00'
    for col_idx, col in enumerate(pay_summary_display.columns, 1):
        ws1.column_dimensions[get_column_letter(col_idx)].width = max(len(str(col)) + 3, 14)
    ws1.freeze_panes = "A2"

    # Sheet 2: Daily Register
    register.to_excel(writer, sheet_name="Daily Payroll Register", index=False)
    ws2 = writer.sheets["Daily Payroll Register"]
    hdr(ws2)
    dt_col = list(register.columns).index("Day Type") + 1
    colors = {
        "Full Day"   : "D5F5E3",
        "Absent"     : "FADBD8",
        "Half Day"   : "FEF9E7",
        "Overtime"   : "D6EAF8",
    }
    for row_idx in range(2, len(register)+2):
        day_type = str(ws2.cell(row=row_idx, column=dt_col).value or "")
        fill_color = next((v for k,v in colors.items() if day_type.startswith(k)), "FFFFFF")
        for col_idx in range(1, len(register.columns)+1):
            cell = ws2.cell(row=row_idx, column=col_idx)
            cell.fill = PatternFill("solid", start_color=fill_color)
            cell.font = Font(name="Calibri", size=9)
            cell.border = thin_border()
            cell.alignment = Alignment(horizontal="center")
    for col_idx, col in enumerate(register.columns, 1):
        ws2.column_dimensions[get_column_letter(col_idx)].width = max(len(str(col)) + 2, 12)
    ws2.freeze_panes = "A2"

    # Sheet 3: Monthly Overview
    monthly.to_excel(writer, sheet_name="Monthly Overview", index=False)
    ws3 = writer.sheets["Monthly Overview"]
    hdr(ws3, color="2E4057")
    for row_idx in range(2, len(monthly)+2):
        for col_idx in range(1, len(monthly.columns)+1):
            cell = ws3.cell(row=row_idx, column=col_idx)
            cell.fill = PatternFill("solid", start_color="EBF5FB" if row_idx%2==0 else "FFFFFF")
            cell.font = Font(name="Calibri", size=10)
            cell.border = thin_border()
            cell.alignment = Alignment(horizontal="center")
    for col_idx, col in enumerate(monthly.columns, 1):
        ws3.column_dimensions[get_column_letter(col_idx)].width = max(len(str(col)) + 2, 14)
    ws3.freeze_panes = "A2"

print(f"\n[DONE] DONE")
print(f"   Total workers:    {len(pay_summary)}")
print(f"   Total hours:      {pay_summary['total_hours'].sum():,.1f}")
print(f"   Total overtime h: {pay_summary['overtime_hours'].sum():,.1f}")
print(f"   Total payroll:    {CURRENCY} {pay_summary['total_pay'].sum():,.2f}")
