"""
SCRIPT 5 — AUTOMATED REPORTING & CHARTS
================================================
What it does:
  - Reads cleaned_data.csv + QC + Payroll outputs
  - Generates a polished multi-sheet Excel report
  - Embeds bar charts: hours worked, validity rates, late arrivals
  - Builds a daily/weekly/monthly breakdown
  - Creates an executive summary sheet with KPIs

Input : cleaned_data.csv
Output: full_report.xlsx
"""

import pandas as pd
import numpy as np
import matplotlib
matplotlib.use("Agg")
import matplotlib.pyplot as plt
import matplotlib.patches as mpatches
import io
from datetime import datetime
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.drawing.image import Image as XLImage


import sys

# Fix Windows console encoding so emojis don't crash the script
try:
    sys.stdout.reconfigure(encoding="utf-8")
except:
    pass

# ── CONFIG ──────────────────────────────────────────────────────────────
INPUT_FILE      = "cleaned_data.csv"
OUTPUT_FILE     = "full_report.xlsx"
ACCURACY_THRESH = 10.0
POINTS_REQUIRED = 10
BOUNDARY = {"lat_min":-1.35,"lat_max":-1.25,"lon_min":36.80,"lon_max":36.85}

# ── HELPERS ──────────────────────────────────────────────────────────────
def thin_border():
    s = Side(style="thin")
    return Border(left=s, right=s, top=s, bottom=s)

def hdr(ws, row=1, color="1F4E79"):
    for cell in ws[row]:
        cell.font = Font(bold=True, color="FFFFFF", name="Calibri", size=11)
        cell.fill = PatternFill("solid", start_color=color)
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = thin_border()
    ws.row_dimensions[row].height = 38

def style_rows(ws, nrows, ncols, alt_color="EBF5FB"):
    for row_idx in range(2, nrows+2):
        for col_idx in range(1, ncols+1):
            cell = ws.cell(row=row_idx, column=col_idx)
            cell.fill = PatternFill("solid", start_color=alt_color if row_idx%2==0 else "FFFFFF")
            cell.font = Font(name="Calibri", size=10)
            cell.border = thin_border()
            cell.alignment = Alignment(horizontal="center")

def auto_width(ws, df):
    for i, col in enumerate(df.columns, 1):
        ws.column_dimensions[get_column_letter(i)].width = min(max(len(str(col))+3, 12), 30)

def parse_time(t):
    for fmt in ["%H:%M:%S","%H:%M"]:
        try:
            return datetime.strptime(str(t).strip(), fmt)
        except: pass
    return None

def chart_to_image(fig):
    buf = io.BytesIO()
    fig.savefig(buf, format="png", dpi=150, bbox_inches="tight")
    buf.seek(0)
    plt.close(fig)
    return buf

# ── 1. LOAD & PREP ───────────────────────────────────────────────────────
print("[LOAD] Loading data...")
df = pd.read_csv(INPUT_FILE)
df["date"]      = pd.to_datetime(df["date"]).dt.date
df["latitude"]  = pd.to_numeric(df["latitude"],  errors="coerce")
df["longitude"] = pd.to_numeric(df["longitude"], errors="coerce")
df["accuracy_m"]= pd.to_numeric(df["accuracy_m"],errors="coerce")
df["week"]      = pd.to_datetime(df["date"].astype(str)).dt.strftime("W%W")
df["month"]     = pd.to_datetime(df["date"].astype(str)).dt.strftime("%Y-%m")

# Validity flags
df["valid"] = (
    df["accuracy_m"].notna() & (df["accuracy_m"] < ACCURACY_THRESH) &
    df["latitude"].between(BOUNDARY["lat_min"], BOUNDARY["lat_max"]) &
    df["longitude"].between(BOUNDARY["lon_min"], BOUNDARY["lon_max"])
)

# Daily per worker
daily = df.drop_duplicates(subset=["worker_id","date"]).copy()
daily["ci"]       = daily["clock_in_time"].apply(parse_time)
daily["co"]       = daily["clock_out_time"].apply(parse_time)
daily["hours"]    = daily.apply(
    lambda r: round((r["co"]-r["ci"]).seconds/3600, 2) if r["ci"] and r["co"] else 0.0, axis=1)
daily["late_min"] = daily["ci"].apply(
    lambda t: max(0, (t.hour*60+t.minute)-(8*60)) if t else 0)

# ── 2. KPI CALCULATIONS ──────────────────────────────────────────────────
total_workers   = df["worker_id"].nunique()
total_points    = len(df)
valid_points    = df["valid"].sum()
validity_pct    = round(valid_points / total_points * 100, 1)
total_hours     = round(daily["hours"].sum(), 1)
avg_hours_day   = round(daily["hours"].mean(), 2)
late_days       = (daily["late_min"] > 5).sum()
absent_pct      = round(100 - (len(daily) / (total_workers * df["date"].nunique()) * 100), 1)

# ── 3. SUMMARY TABLES ────────────────────────────────────────────────────
worker_summary = df.groupby(["worker_id","worker_name","team"]).agg(
    total_points  = ("valid",   "count"),
    valid_points  = ("valid",   "sum"),
    days_active   = ("date",    "nunique"),
).reset_index()
worker_summary["validity_pct"] = round(worker_summary["valid_points"]/worker_summary["total_points"]*100,1)

hours_summary = daily.groupby(["worker_id","worker_name"]).agg(
    total_hours = ("hours",    "sum"),
    avg_hours   = ("hours",    "mean"),
    late_days   = ("late_min", lambda x: (x > 5).sum()),
).reset_index().round(2)

weekly_summary = df.groupby("week").agg(
    total_points  = ("valid",       "count"),
    valid_points  = ("valid",       "sum"),
    workers_active= ("worker_id",   "nunique"),
).reset_index()
weekly_summary["validity_pct"] = round(weekly_summary["valid_points"]/weekly_summary["total_points"]*100,1)

# ── 4. CHARTS ────────────────────────────────────────────────────────────
print(" Generating charts...")
COLORS = ["#2196F3","#4CAF50","#FF5722","#9C27B0","#FF9800",
          "#00BCD4","#E91E63","#8BC34A","#FFC107","#3F51B5"]

# Chart A: Total hours per worker
fig_a, ax = plt.subplots(figsize=(10, 5))
names = hours_summary["worker_name"].str.split().str[0]
ax.bar(names, hours_summary["total_hours"], color=COLORS[:len(names)])
ax.set_title("Total Hours Worked per Worker", fontsize=14, fontweight="bold")
ax.set_xlabel("Worker"); ax.set_ylabel("Hours")
ax.tick_params(axis="x", rotation=30)
plt.tight_layout()
img_hours = chart_to_image(fig_a)

# Chart B: Validity % per worker
fig_b, ax = plt.subplots(figsize=(10, 5))
colors_b = ["#4CAF50" if v >= 95 else ("#FFC107" if v >= 80 else "#F44336")
            for v in worker_summary["validity_pct"]]
ax.barh(worker_summary["worker_name"].str.split().str[0],
        worker_summary["validity_pct"], color=colors_b)
ax.axvline(95, color="green",  linestyle="--", linewidth=1.5, label="PASS (95%)")
ax.axvline(80, color="orange", linestyle="--", linewidth=1.5, label="REVIEW (80%)")
ax.set_title("GPS Data Validity % per Worker", fontsize=14, fontweight="bold")
ax.set_xlabel("Validity %"); ax.set_xlim(0, 105)
ax.legend(loc="lower right")
plt.tight_layout()
img_valid = chart_to_image(fig_b)

# Chart C: Late days per worker
fig_c, ax = plt.subplots(figsize=(10, 5))
ax.bar(hours_summary["worker_name"].str.split().str[0],
       hours_summary["late_days"],
       color=["#F44336" if v > 3 else "#FF9800" if v > 1 else "#4CAF50"
              for v in hours_summary["late_days"]])
ax.axhline(3, color="red", linestyle="--", linewidth=1, label="Chronic Late threshold")
ax.set_title("Late Clock-Ins per Worker", fontsize=14, fontweight="bold")
ax.set_xlabel("Worker"); ax.set_ylabel("Days Late")
ax.legend(); ax.tick_params(axis="x", rotation=30)
plt.tight_layout()
img_late = chart_to_image(fig_c)

# Chart D: Weekly validity trend
fig_d, ax = plt.subplots(figsize=(10, 4))
ax.plot(weekly_summary["week"], weekly_summary["validity_pct"],
        marker="o", color="#2196F3", linewidth=2, markersize=8)
ax.fill_between(range(len(weekly_summary)), weekly_summary["validity_pct"],
                alpha=0.15, color="#2196F3")
ax.set_title("Weekly Data Validity Trend", fontsize=14, fontweight="bold")
ax.set_xlabel("Week"); ax.set_ylabel("Validity %")
ax.set_xticks(range(len(weekly_summary))); ax.set_xticklabels(weekly_summary["week"])
ax.set_ylim(0, 105)
plt.tight_layout()
img_trend = chart_to_image(fig_d)

# ── 5. WRITE EXCEL ───────────────────────────────────────────────────────
print(f"[SAVE] Writing {OUTPUT_FILE}...")

with pd.ExcelWriter(OUTPUT_FILE, engine="openpyxl") as writer:

    # ── Executive Summary sheet ──────────────────────────────────────────
    ws_exec = writer.book.create_sheet("Executive Summary")

    # Title
    ws_exec.merge_cells("A1:H1")
    title = ws_exec["A1"]
    title.value = " FIELD DATA COLLECTION — EXECUTIVE SUMMARY"
    title.font  = Font(bold=True, size=16, name="Calibri", color="FFFFFF")
    title.fill  = PatternFill("solid", start_color="1F4E79")
    title.alignment = Alignment(horizontal="center", vertical="center")
    ws_exec.row_dimensions[1].height = 45

    # KPI Cards
    kpis = [
        ("Total Workers",         total_workers,   "2196F3"),
        ("Total Points Collected",total_points,    "4CAF50"),
        ("Valid Points",          int(valid_points),"66BB6A"),
        ("Overall Validity %",    f"{validity_pct}%","FFC107"),
        ("Total Hours Worked",    f"{total_hours}h","9C27B0"),
        ("Avg Hours / Day",       f"{avg_hours_day}h","FF9800"),
        ("Late Clock-ins",        int(late_days),  "F44336"),
        ("Absence Rate",          f"{absent_pct}%","78909C"),
    ]

    ws_exec.row_dimensions[3].height = 20
    for i, (label, value, color) in enumerate(kpis):
        col_start = (i % 4) * 2 + 1
        row_start = 4 if i < 4 else 7

        lbl_cell = ws_exec.cell(row=row_start,   column=col_start, value=label)
        val_cell = ws_exec.cell(row=row_start+1, column=col_start, value=value)

        for cell, sz, bold in [(lbl_cell, 9, False), (val_cell, 18, True)]:
            cell.font      = Font(bold=bold, size=sz, name="Calibri",
                                  color="FFFFFF" if bold else "DDDDDD")
            cell.fill      = PatternFill("solid", start_color=color)
            cell.alignment = Alignment(horizontal="center", vertical="center")
            cell.border    = thin_border()
        ws_exec.row_dimensions[row_start+1].height = 35
        ws_exec.column_dimensions[get_column_letter(col_start)].width = 22

    # Add charts
    for buf, anchor in [(img_hours,"A10"),(img_valid,"E10"),
                        (img_late,"A29"),(img_trend,"E29")]:
        img = XLImage(buf)
        img.width  = 480
        img.height = 240
        ws_exec.add_image(img, anchor)

    # ── Worker Summary sheet ─────────────────────────────────────────────
    worker_summary.to_excel(writer, sheet_name="Worker Summary", index=False)
    ws2 = writer.sheets["Worker Summary"]
    hdr(ws2)
    style_rows(ws2, len(worker_summary), len(worker_summary.columns))
    auto_width(ws2, worker_summary)
    ws2.freeze_panes = "A2"

    # ── Hours Summary sheet ──────────────────────────────────────────────
    hours_summary.to_excel(writer, sheet_name="Hours Summary", index=False)
    ws3 = writer.sheets["Hours Summary"]
    hdr(ws3, color="2E4057")
    style_rows(ws3, len(hours_summary), len(hours_summary.columns))
    auto_width(ws3, hours_summary)
    ws3.freeze_panes = "A2"

    # ── Weekly Trend sheet ───────────────────────────────────────────────
    weekly_summary.to_excel(writer, sheet_name="Weekly Trend", index=False)
    ws4 = writer.sheets["Weekly Trend"]
    hdr(ws4, color="5D3A9B")
    style_rows(ws4, len(weekly_summary), len(weekly_summary.columns))
    auto_width(ws4, weekly_summary)

    # Set exec summary as first sheet
    writer.book.move_sheet("Executive Summary", offset=-len(writer.book.sheetnames)+1)

print(f"\n[DONE] DONE — {OUTPUT_FILE}")
print(f"\n    Key Stats:")
print(f"      Workers:     {total_workers}")
print(f"      Points:      {total_points}")
print(f"      Validity:    {validity_pct}%")
print(f"      Total hours: {total_hours}h")
